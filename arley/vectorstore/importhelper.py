# nomic-embed-text:latest
import argparse
import glob
import re
import sys
from enum import StrEnum, auto
from functools import partial
from io import StringIO
from pathlib import Path, PurePath
from typing import (Any, AnyStr, Dict, Generator, Iterator, List, Literal,
                    Mapping, Optional, Sequence, Tuple)

import chromadb
import llama_index
import openparse
import openpyxl
import openpyxl.utils.cell
import ruamel.yaml
from chromadb.api.models.Collection import Collection as ChromaCollection
from llama_index.core.node_parser import (
    LanguageConfig, MarkdownNodeParser,
    SemanticDoubleMergingSplitterNodeParser, SemanticSplitterNodeParser,
    SentenceSplitter)
from llama_index.core.schema import BaseNode, MetadataMode, TextNode
from llama_index.embeddings.ollama import OllamaEmbedding
from llama_index.readers.file import FlatReader
from llama_index.readers.file.markdown import MarkdownReader
from loguru import logger
from mammoth.results import Result
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from arley import Helper
from arley.config import (CHROMADB_DEFAULT_COLLECTION_NAME, OLLAMA_EMBED_MODEL,
                          OLLAMA_GUESS_LANGUAGE_MODEL, OLLAMA_HOST,
                          OLLAMA_MODEL, OllamaPrimingMessage, TemplateType,
                          get_ollama_options, is_in_cluster, settings)
from arley.dbobjects.ragdoc import ArleyDocumentInformation, DocTypeEnum
from arley.llm import ollama_adapter
from arley.llm.language_guesser import LanguageGuesser
from arley.llm.ollama_adapter import (
    Message, ask_ollama_chat, get_create_outline_prompt,
    get_create_summary_prompt, get_reformat_and_semantically_markup_prompt)
from arley.vectorstore.chroma_adapter import ChromaDBConnection

# import yaml










_OLLAMA_DEFAULT_MODEL: str = (
    OLLAMA_MODEL  # "nous-hermes2-mixtral:8x7b"  # "hermes3:8b-llama3.1-fp16"  # "hermes3:70b-llama3.1-q4_0"  # "nous-hermes2-mixtral:8x7b"
)
_OLLAMA_DEFAULT_EMBED_MODEL: str = (
    OLLAMA_EMBED_MODEL  # "nomic-embed-text:latest"  #"mxbai-embed-large:latest"  # "nomic-embed-text:latest"
)
_CHROMADB_DEFAULT_COLLECTION_NAME: str = CHROMADB_DEFAULT_COLLECTION_NAME

# import numpy as np


def docx_to_html(file: Path) -> Path | None:
    # https://github.com/mwilliamson/python-mammoth
    import mammoth

    with open(file, "rb") as docx_file:
        result: Result = mammoth.convert_to_html(docx_file)
        html = result.value  # The generated HTML
        messages = result.messages  # Any messages, such as warnings during conversion

        outf = Path(file.parent.resolve(), file.name[: file.name.find(".doc")] + ".html")

        # with open(outf, 'w') as html_file:
        outf.write_text(html)
        logger.debug(f"written to: {outf.absolute()}")

        # You can also extract the raw text of the document by using mammoth.extract_raw_text. This will ignore all formatting in the document. Each paragraph is followed by two newlines.
        # with open("document.docx", "rb") as docx_file:
        #     result = mammoth.extract_raw_text(docx_file)
        #     text = result.value # The raw text
        #     messages = result.messages # Any messages

        return outf


def html_to_markdown(file: Path, tags_to_strip: None | list[str] = None) -> Path | None:
    from markdownify import markdownify as md

    # md('<b>Yay</b> <a href="http://github.com">GitHub</a>', convert=['b'])  # > '**Yay** GitHub'

    html_snippet: str = file.read_text()

    markdowned: str = md(html_snippet, strip=tags_to_strip)

    outf = Path(file.parent.resolve(), file.name[: file.name.rfind(MyFileTypes.html.value)] + MyFileTypes.md.value)
    # with open(outf, 'w') as html_file:
    outf.write_text(markdowned)
    logger.debug(f"written to: {outf.absolute()}")

    return outf


def xlsx_to_yaml_raw(file: Path) -> Path | None:
    outf: Path = Path(file.parent.resolve(), file.name[0 : file.name.rfind(".")] + "_raw." + MyFileTypes.yaml.value)
    if outf.exists():
        logger.debug(f"EXISTS: {outf.resolve()}")
        return outf
    outf.touch()

    colname_to_idx: Dict[str, int] = {}
    idx_to_colname: Dict[int, str] = {}

    xlsx_data: Dict[str, List[str | int] | None] = {}

    # OR read with pandas: https://pythonbasics.org/read-excel/

    # Define variable to load the dataframe
    wb: Workbook = openpyxl.load_workbook(file)

    cell: openpyxl.cell.cell.Cell
    # Define variable to read sheet
    sheetnames: list[str] = wb.sheetnames
    for sheetname in sheetnames:
        sheet: Worksheet = wb[sheetname]
        logger.debug(f"{sheetname=} {type(sheet)=} {sheet=}")
        logger.debug(f"{sheet.max_column=} {sheet.max_row=}")

        sheet.delete_rows(idx=0, amount=3)
        # sheetname='NDA_1' type(sheet)=<class 'openpyxl.worksheet.worksheet.Worksheet'> sheet=<Worksheet "NDA_1">
        logger.debug(f"{sheet.max_column=} {sheet.max_row=}")

        for rowidx, row in enumerate(sheet.iter_rows()):  # min_row=1, max_col=3, max_row=2):
            logger.debug(f"{rowidx=} {type(row)=} {row=}")

            for cellidx, cell in enumerate(row):
                # logger.debug(f"{type(cell)=} {cell=}")
                # logger.debug(f"{cell.col_idx=} {cell.coordinate=} {cell.value=}")
                # logger.debug(f"{openpyxl.utils.cell.coordinate_to_tuple(cell.coordinate)=}")

                if rowidx == 0:
                    colname_to_idx[cell.value] = cellidx
                    idx_to_colname[cellidx] = cell.value

                    xlsx_data[cell.value] = []  # type: ignore

                    continue

                if cell.value:
                    value: Any = cell.value
                    if idx_to_colname[cellidx] == "Relevanz":
                        value = value * 100
                    elif isinstance(value, str):
                        value = value.strip()

                    xlsx_data[idx_to_colname[cellidx]].append(value)  # type: ignore

    for key in xlsx_data:  # .keys():
        values: Any = xlsx_data[key]
        if isinstance(values, list):
            if len(values) == 1:
                xlsx_data[key] = values[0]
            elif len(values) == 0:
                xlsx_data[key] = None

    yaml = ruamel.yaml.YAML()
    yaml.indent(sequence=4, offset=2)

    with open(outf, "w") as outfile:
        # yaml.dump(xlsx_data, outfile, indent=4, allow_unicode=True, sort_keys=False, default_flow_style=False)
        yaml.dump(xlsx_data, outfile)
        # yaml.dump(xlsx_data, sys.stdout)
        logger.debug(Helper.get_dict_as_yaml_str(xlsx_data))

    return outf


def docx_to_html_via_subprocess(file: Path) -> Tuple[int, Path] | None:
    outf: Path = Path(
        file.parent.resolve(), file.name[: file.name.rfind(MyFileTypes.docx.value)] + MyFileTypes.html.value
    )
    cwd: str = str(file.parent.resolve().absolute())
    cmd_array: list[str] = ["/usr/bin/libreoffice", "--invisible", "--convert-to", "html", str(file.absolute())]

    ret: int = exec_sub_process(cwd=cwd, cmd_array=cmd_array)

    return ret, outf


def docx_to_md_via_subprocess(file: Path) -> Tuple[int, Path] | None:
    outf_md = Path(file.parent.resolve(), file.name[: file.name.rfind(MyFileTypes.docx.value)] + MyFileTypes.md.value)
    cwd: str = str(file.parent.resolve().absolute())

    # markdown_mmd markdown markdown_strict gfm commonmark
    cmd_array: list[str] = [
        "/usr/bin/pandoc",
        "-f",
        "docx",
        "-t",
        "commonmark",
        "-o",
        str(outf_md.absolute()),
        str(file.absolute()),
    ]

    ret: int = exec_sub_process(cwd=cwd, cmd_array=cmd_array)

    return ret, outf_md


def docx_to_txt_via_subprocess(file: Path) -> Tuple[int, Path] | None:
    outf_txt = Path(file.parent.resolve(), file.name[: file.name.rfind(MyFileTypes.docx.value)] + MyFileTypes.txt.value)
    cwd: str = str(file.parent.resolve().absolute())

    # markdown_mmd markdown markdown_strict gfm commonmark
    cmd_array: list[str] = [
        "/usr/bin/pandoc",
        "-f",
        "docx",
        "-t",
        "plain",
        "-o",
        str(outf_txt.absolute()),
        str(file.absolute()),
    ]

    ret: int = exec_sub_process(cwd=cwd, cmd_array=cmd_array)

    return ret, outf_txt


def docx_to_pdf_via_subprocess(file: Path) -> Tuple[int, Path] | None:
    outf: Path = Path(
        file.parent.resolve(), file.name[: file.name.rfind(MyFileTypes.docx.value)] + MyFileTypes.pdf.value
    )
    cwd: str = str(file.parent.resolve().absolute())
    cmd_array: list[str] = ["/usr/bin/libreoffice", "--invisible", "--convert-to", "pdf", str(file.absolute())]

    ret: int = exec_sub_process(cwd=cwd, cmd_array=cmd_array)

    return ret, outf


def html_to_md_via_subprocess(file: Path) -> Tuple[int, Path] | None:
    outf_md = Path(file.parent.resolve(), file.name[: file.name.rfind(MyFileTypes.html.value)] + MyFileTypes.md.value)
    cwd: str = str(file.parent.resolve().absolute())

    cmd_array: list[str] = [
        "/usr/bin/pandoc",
        "-f",
        "html",
        "-t",
        "markdown_strict",
        "-o",
        str(outf_md.absolute()),
        str(file.absolute()),
    ]

    ret: int = exec_sub_process(cwd=cwd, cmd_array=cmd_array)

    return ret, outf_md


def exec_sub_process(cwd: str, cmd_array: list[str]) -> int:
    # BASED ON: https://gist.githubusercontent.com/goerz/8897c2d8a602af2a45d4/raw/d7bf5a5d589389e99c796a555b640f8f4268d93e/doc2markdown.py

    import subprocess
    import sys

    ret: int = subprocess.call(cmd_array, stderr=sys.stderr, stdout=sys.stdout, cwd=cwd)

    return ret


# def cosine_similarity(
#     a: np.ndarray | list[float], b: np.ndarray | list[float]
# ) -> float:
#     return np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b))
#
#
# def get_node_similarities(nodes: list[Node]):
#     # get the similarity of each node with the node that precedes it
#     embeddings = embedding_client.embed_many([node.text for node in nodes])
#     similarities = []
#     for i in range(1, len(embeddings)):
#         similarities.append(cosine_similarity(embeddings[i - 1], embeddings[i]))
#
#     similarities = [round(sim, 2) for sim in similarities]
#     return [0] + similarities


def parse_pdf(file: Path) -> int:
    parser: openparse.DocumentParser = openparse.DocumentParser()
    parsed_basic_doc: openparse.schemas.ParsedDocument = parser.parse(file)

    for node in parsed_basic_doc.nodes:
        node_data: dict = node.model_dump(mode="json", by_alias=True, exclude_none=True)
        logger.debug(Helper.get_pretty_dict_json_no_sort(node_data))

    # lala = openparse.processing.SemanticIngestionPipeline()
    # from openparse import processing, DocumentParser
    #
    # semantic_pipeline = processing.SemanticIngestionPipeline(
    #     openai_api_key=OPEN_AI_KEY,
    #     model="text-embedding-3-large",
    #     min_tokens=64,
    #     max_tokens=1024,
    # )
    # parser = DocumentParser(
    #     processing_pipeline=semantic_pipeline,
    # )
    # parsed_content = parser.parse(basic_doc_path)

    return 0


def llm_get_helpfull_atorney_system_prompt(
    lang: Literal["en", "de"] = "en", return_in_markdown: bool = True
) -> str | None:
    priming_msg: OllamaPrimingMessage
    for priming_msg in settings.ollama.ollama_priming_messages:
        if priming_msg.lang != lang or priming_msg.role != "system":
            continue

        ret: str = priming_msg.content
        if return_in_markdown:
            if lang == "en":
                ret = ret.rstrip() + "\nYour output must be formatted in markdown syntax."
            else:
                ret = ret.rstrip() + "\nDeine Ausgabe muß in markdown syntax formatiert sein."
        return ret

    return None


def llm_remove_unnecessary_newlines_and_generate_semantic_markup_file(
    txt_file: Path,
    ollama_model: str = _OLLAMA_DEFAULT_MODEL,
    temperature: float = 0.0,
    template_type: TemplateType = TemplateType.xml,
) -> Path | None:
    outf: Path = Path(
        txt_file.parent.resolve(), txt_file.name[: txt_file.name.rfind(".")] + "_llm." + MyFileTypes.md.value
    )
    outf.touch()

    _txt: str
    with open(txt_file, "r") as infile:
        _txt = infile.read()

    new_text = llm_remove_unnecessary_newlines_and_generate_semantic_markup(
        plain_txt=_txt, ollama_model=ollama_model, template_type=template_type, temperature=temperature
    )

    assert new_text is not None

    with open(outf, "w") as outfile:
        outfile.write(new_text)
        outfile.flush()

    logger.debug(new_text)

    return outf


# "nous-hermes2-mixtral:8x7b"  # "llama3.1:70b-instruct-q3_K_M"  # "llama3.1:latest"  # "llama3.1:70b-instruct-q4_0"  #"nous-hermes2-mixtral:8x7b"  #"gemma2:27b"  #mixtral:latest"
def llm_remove_unnecessary_newlines_and_generate_semantic_markup(
    plain_txt: str,
    ollama_model: str = _OLLAMA_DEFAULT_MODEL,
    temperature: float = 0.0,
    template_type: TemplateType = TemplateType.xml,
) -> str | None:
    ret: str | None

    lang: Literal["en", "de"] = guess_my_language(txt=plain_txt, ollama_model=ollama_model)

    msgs: list[Message] | None = None  # []

    prompt: str = get_reformat_and_semantically_markup_prompt(
        md_or_plaintext=plain_txt, lang=lang, ollama_model=ollama_model, template_type=template_type
    )

    logger.debug(prompt)

    resp: Dict[Any, Any] | Mapping[str, Any] | Iterator[Mapping[str, Any]] = ask_ollama_chat(
        streamed=True,
        system_prompt=None,
        prompt=prompt,
        msg_history=msgs,
        model=ollama_model,
        evict=False,
        temperature=temperature,
        top_k=40,
        top_p=0.9,
        num_predict=-1,
        seed=0,
        print_msgs=False,
        print_response=True,
    )

    assert isinstance(resp, dict)

    new_text: str = resp["message"]["content"]
    ret = new_text

    assert isinstance(ret, str)

    return ret


def llm_create_outline_file(
    md_or_plaintxt_file: Path,
    is_plaintext: bool,
    ollama_model: str = _OLLAMA_DEFAULT_MODEL,
    delete_other_llm_base: bool = True,
    temperature: float = 0.4,
    template_type: TemplateType = TemplateType.xml,
) -> Path | None:
    outf_name: str = md_or_plaintxt_file.name[: md_or_plaintxt_file.name.rfind(".")] + "_outline."
    outf_name_other: str = outf_name
    if is_plaintext:
        outf_name += MyFileTypes.txt.value
        outf_name_other += MyFileTypes.md.value
    else:
        outf_name += MyFileTypes.md.value
        outf_name_other += MyFileTypes.txt.value

    outf: Path = Path(md_or_plaintxt_file.parent.resolve(), outf_name)
    outf.touch()

    outf_other: Path = Path(md_or_plaintxt_file.parent.resolve(), outf_name_other)

    if delete_other_llm_base:
        outf_other.unlink(missing_ok=True)

    md_or_plaintxt: str
    with open(md_or_plaintxt_file, "r") as infile:
        md_or_plaintxt = infile.read()

    outline_text = llm_create_outline(
        md_or_plaintxt=md_or_plaintxt,
        is_plaintext=is_plaintext,
        ollama_model=ollama_model,
        template_type=template_type,
        temperature=temperature,
    )

    assert outline_text is not None

    with open(outf, "w") as outfile:
        outfile.write(outline_text)
        outfile.flush()

    logger.debug(outline_text)

    return outf


def guess_my_language(
    txt: str,
    print_response: bool = True,
    ollama_model: str = OLLAMA_GUESS_LANGUAGE_MODEL,
    print_msgs: bool = True,
    print_detect_txt: bool = True,
) -> Literal["en", "de"]:
    lang_detect_text: str = f"{txt[:min(len(txt), 512)]}".strip()

    lang: Literal["en", "de"] | None = None
    ollama_response: dict | None = None
    lang_detect_content: dict | None = None

    if print_detect_txt:
        logger.debug(f"\n{lang_detect_text=}\n")

    ret: Literal["de", "en"] | Tuple[Literal["de", "en"], dict, dict] | None = None
    try:
        ret = LanguageGuesser.guess_language(
            input_text=lang_detect_text,
            only_return_str=False,
            ollama_host=OLLAMA_HOST,
            ollama_model=ollama_model,
            ollama_options=get_ollama_options(ollama_model),
            print_msgs=print_msgs,
            print_response=print_response,
            print_http_response=False,
            print_http_request=False,
            max_retries=3,
        )
    except Exception as ex:
        logger.exception(Helper.get_exception_tb_as_string(ex))

    if ret and isinstance(ret, tuple):
        lang, ollama_response, lang_detect_content = ret

    if lang is None:
        lang = "de"
        logger.error("DEFAULTING LANG to de")

    return lang  # type: ignore


# _OLLAMA_DEFAULT_MODEL  # "llama3.1:70b-instruct-q3_K_M"  # "llama3.1:latest"  # "llama3.1:70b-instruct-q4_0"  #_OLLAMA_DEFAULT_MODEL  #"gemma2:27b"  #mixtral:latest"
def llm_create_outline(
    md_or_plaintxt: str,
    is_plaintext: bool,
    ollama_model: str = _OLLAMA_DEFAULT_MODEL,
    temperature: float = 0.4,
    template_type: TemplateType = TemplateType.xml,
) -> str | None:
    ret: str | None = None

    lang: Literal["en", "de"] = guess_my_language(txt=md_or_plaintxt, ollama_model=ollama_model)

    msgs: list[Message] | None = None  # []

    system_prompt: str | None = llm_get_helpfull_atorney_system_prompt(
        lang="en", return_in_markdown=False if is_plaintext else True
    )  # always en!

    prompt: str = get_create_outline_prompt(
        md_or_plaintext=md_or_plaintxt, lang=lang, ollama_model=ollama_model, template_type=template_type
    )

    logger.debug(system_prompt)
    logger.debug(prompt)

    resp: Dict[Any, Any] | Mapping[str, Any] | Iterator[Mapping[str, Any]] = ask_ollama_chat(
        streamed=True,
        system_prompt=system_prompt,
        prompt=prompt,
        msg_history=msgs,
        model=ollama_model,
        evict=False,
        temperature=temperature,
        top_k=40,
        top_p=0.9,
        num_predict=-1,
        seed=0,
        print_msgs=False,
        print_response=True,
    )

    if isinstance(resp, dict):
        new_text: str = resp["message"]["content"]
        ret = new_text

    return ret


def llm_create_summary_file(
    md_or_plaintxt_file: Path,
    is_plaintext: bool,
    ollama_model: str = _OLLAMA_DEFAULT_MODEL,
    delete_other_llm_base: bool = True,
    temperature: float = 0.4,
    template_type: TemplateType = TemplateType.xml,
) -> Path | None:
    outf_name: str = md_or_plaintxt_file.name[: md_or_plaintxt_file.name.rfind(".")] + "_summary."
    outf_name_other: str = outf_name
    if is_plaintext:
        outf_name += MyFileTypes.txt.value
        outf_name_other += MyFileTypes.md.value
    else:
        outf_name += MyFileTypes.md.value
        outf_name_other += MyFileTypes.txt.value

    outf: Path = Path(md_or_plaintxt_file.parent.resolve(), outf_name)
    outf.touch()

    outf_other: Path = Path(md_or_plaintxt_file.parent.resolve(), outf_name_other)

    if delete_other_llm_base:
        outf_other.unlink(missing_ok=True)

    md_or_plaintxt: str
    with open(md_or_plaintxt_file, "r") as infile:
        md_or_plaintxt = infile.read()

    summary_text: str | None = llm_create_summary(
        md_or_plaintxt=md_or_plaintxt,
        is_plaintext=is_plaintext,
        ollama_model=ollama_model,
        temperature=temperature,
        template_type=template_type,
    )

    assert summary_text is not None

    with open(outf, "w") as outfile:
        outfile.write(summary_text)
        outfile.flush()

    logger.debug(summary_text)

    return outf


# _OLLAMA_DEFAULT_MODEL  # "llama3.1:70b-instruct-q3_K_M"  # "llama3.1:latest"  # "llama3.1:70b-instruct-q4_0"  #_OLLAMA_DEFAULT_MODEL  #"gemma2:27b"  #mixtral:latest"
def llm_create_summary(
    md_or_plaintxt: str,
    is_plaintext: bool,
    ollama_model: str = _OLLAMA_DEFAULT_MODEL,
    temperature: float = 0.4,
    template_type: TemplateType = TemplateType.xml,
) -> str | None:
    ret: str | None = None

    lang: Literal["en", "de"] = guess_my_language(txt=md_or_plaintxt, ollama_model=ollama_model)

    msgs: list[Message] | None = None  # []

    system_prompt: str | None = llm_get_helpfull_atorney_system_prompt(
        lang="en", return_in_markdown=False if is_plaintext else True
    )

    prompt: str = get_create_summary_prompt(
        md_or_plaintext=md_or_plaintxt, lang=lang, ollama_model=ollama_model, template_type=template_type
    )

    logger.debug(system_prompt)
    logger.debug(prompt)

    resp: Dict[Any, Any] | Mapping[str, Any] | Iterator[Mapping[str, Any]] = ask_ollama_chat(
        streamed=True,
        system_prompt=system_prompt,
        prompt=prompt,
        msg_history=msgs,
        model=ollama_model,
        evict=False,
        temperature=temperature,
        top_k=40,
        top_p=0.9,
        num_predict=-1,
        seed=0,
        print_msgs=False,
        print_response=True,
    )

    assert isinstance(resp, dict)

    ret = resp["message"]["content"]

    return ret


def create_excerpt_files(
    md_or_plaintxt_file: Path,
    is_plaintext: bool,
    spacy_mode: bool = False,
    do_simple_md_split: bool = True,
    lang: Optional[Literal["en", "de"]] = None,
    delete_old_excerpts: bool = True,
) -> list[Path] | None:
    out_dir: Path = Path(md_or_plaintxt_file.parent.resolve(), "llm_excerpts")
    out_dir.mkdir(exist_ok=True)
    existing_files: Iterator[Path] = out_dir.glob("*")
    for f in existing_files:
        if delete_old_excerpts:
            logger.debug(f"Already exists (and will be deleted): ${f.resolve()}")
            f.unlink()
        else:
            logger.debug(f"Already exists (and will be kept): ${f.resolve()}")

    ret: list[Path] = []

    nodes: List[TextNode | BaseNode] = parse_markdown_or_plaintext_semantically(
        md_or_plaintext_file=md_or_plaintxt_file,
        is_plaintext=is_plaintext,
        spacy_mode=spacy_mode,
        do_simple_md_split=do_simple_md_split,
        lang=lang,
    )  # ignore lang for simple split

    for node in nodes:
        if not isinstance(node, TextNode):
            continue

        logger.debug(f"{type(node)=} {node.start_char_idx=} {node.end_char_idx=} {node.get_node_info()=}")

        node_data: dict = node.model_dump(mode="json", by_alias=True, exclude_none=True)
        logger.debug(Helper.get_pretty_dict_json_no_sort(node_data))

        fromchar: int | None = node.start_char_idx
        tochar: int | None = node.end_char_idx

        excerpt_name: str = (
            md_or_plaintxt_file.name[: md_or_plaintxt_file.name.rfind(".")]
            + f"_excerpt_{fromchar}-{tochar}."
            + MyFileTypes.txt.value
        )

        txt_out: str = node.get_content().strip()

        if len(txt_out) == 0:
            continue

        out_f: Path = Path(out_dir, excerpt_name)
        with open(out_f, "w") as outfile:
            outfile.write(txt_out)

        ret.append(out_f)

    return ret if len(ret) > 0 else None


def parse_markdown_or_plaintext_semantically(
    md_or_plaintext_file: Path,
    is_plaintext: bool,
    spacy_mode: bool = False,
    do_simple_md_split: bool = True,
    lang: Optional[Literal["en", "de"]] = None,
) -> List[TextNode | BaseNode]:

    md_or_plaintext_docs_flat: list[llama_index.core.schema.Document] = FlatReader().load_data(md_or_plaintext_file)

    if not lang:
        if spacy_mode or not do_simple_md_split:
            lang = guess_my_language(txt=md_or_plaintext_docs_flat[0].get_content())
    logger.debug(f"LANG: {lang}")

    doc_nodes: list[TextNode | BaseNode] = semantic_split(
        is_plaintext=is_plaintext,
        spacy_mode=spacy_mode,
        lang=lang,
        doc=md_or_plaintext_docs_flat[0],
        do_simple_md_split=do_simple_md_split,
    )  # do not pass in already parsed markdown-docs -> no str index and such...
    logger.debug(f"{len(doc_nodes)=}")

    text_out: StringIO = StringIO()
    doc_node: TextNode | BaseNode
    for doc_node in doc_nodes:
        if not isinstance(doc_node, TextNode):
            continue

        logger.debug(
            f"{type(doc_node)=} {doc_node.start_char_idx=} {doc_node.end_char_idx=} {doc_node.get_node_info()=}"
        )

        node_data: dict = doc_node.model_dump(mode="json", by_alias=True, exclude_none=True)
        logger.debug(Helper.get_pretty_dict_json_no_sort(node_data))

        text_out.write("\n#####################\n")
        text_out.write(doc_node.get_content().strip())

    logger.debug(text_out.getvalue())

    return doc_nodes


def parse_markdown(
    file: Path, remove_leading_comment_indent: bool = True, remove_triple_newline: bool = True
) -> str | None:
    # md_docs = MarkdownReader().load_data(file)
    md_docs = FlatReader().load_data(file)
    parser: MarkdownNodeParser = MarkdownNodeParser()

    text_out: StringIO = StringIO()

    doc_nodes: list[BaseNode] = parser.get_nodes_from_documents(md_docs, show_progress=True)
    logger.debug(f"{len(doc_nodes)=}")
    for doc_node in doc_nodes:
        # doc_node_parser: MarkdownNodeParser = MarkdownNodeParser()

        text_nodes: list[TextNode] = parser.get_nodes_from_node(
            doc_node
        )  # doc_node_parser.get_nodes_from_node(doc_node)
        for node in text_nodes:
            # node_data: dict = node.model_dump(mode="json", by_alias=True, exclude_none=True)
            # logger.debug(Helper.get_pretty_dict_json_no_sort(node_data))
            ct: str = node.get_content()
            if remove_leading_comment_indent:
                ct = re.sub(r"^> ", "", ct, flags=re.MULTILINE)
            if remove_triple_newline:
                ct = re.sub(r"^\n\n\n", "\n", ct, flags=re.MULTILINE)
            text_out.write(ct)

    logger.debug(text_out.getvalue())

    return text_out.getvalue()


def semantic_split(
    is_plaintext: bool,
    spacy_mode: bool,
    lang: Optional[Literal["en", "de"]],
    doc: llama_index.core.schema.Document,
    ollama_embed_model: str = _OLLAMA_DEFAULT_EMBED_MODEL,
    do_simple_md_split: bool = True,
) -> List[TextNode | BaseNode]:
    def my_id_func(prefix: str, index: int, document: Any) -> str:
        # logger.debug(f"{index=} {document=}")
        return f"{prefix}-my-new-node-id-{index}"

    # sentence_splitter: SentenceSplitter = SentenceSplitter.from_defaults(
    #     include_metadata=True
    # )
    #
    # def my_sentence_splitter_callable(text: str) -> list[str]:
    #     ret: list[str] =  sentence_splitter.split_text(text)
    #     logger.debug(f"{text} -> {len(ret)=}")
    #     return ret

    # logger.debug(f"{type(doc)=} {doc=}")

    semantic_nodes: list[TextNode | BaseNode]

    if spacy_mode:
        raise RuntimeError("spacy mode is not supported at the moment (HT 20240905).")
        # config: LanguageConfig
        # nltk.download("punkt_tab")
        #
        # if lang == "en":
        #     config = LanguageConfig(language="english", spacy_model="en_core_web_md")
        # else:
        #     config = LanguageConfig(language="german", spacy_model="de_core_news_md")
        #
        # dm_semantic_splitter: SemanticDoubleMergingSplitterNodeParser = SemanticDoubleMergingSplitterNodeParser(
        #     language_config=config,
        #     initial_threshold=0.4,
        #     appending_threshold=0.5,
        #     merging_threshold=0.5,
        #     max_chunk_size=5000,
        #     show_progress=True,
        #     merging_range=2
        # )
        #
        # semantic_nodes = dm_semantic_splitter.get_nodes_from_documents([doc])
    else:
        if do_simple_md_split and not is_plaintext:
            parser: MarkdownNodeParser = MarkdownNodeParser()

            semantic_nodes = parser.get_nodes_from_documents([doc], show_progress=True)  # type: ignore
            logger.debug(f"{len(semantic_nodes)=}")
            # for doc_node in doc_nodes:
            #     text_nodes: list[TextNode] = parser.get_nodes_from_node(doc_node)  # doc_node_parser.get_nodes_from_node(doc_node)
            #     for node in text_nodes:
            #         node_data: dict = node.model_dump(mode="json", by_alias=True, exclude_none=True)
            #         logger.debug(Helper.get_pretty_dict_json_no_sort(node_data))

        else:
            embed_model: OllamaEmbedding = OllamaEmbedding(
                model_name=ollama_embed_model, base_url=ollama_adapter.OLLAMA_HOST
            )

            # breakpoint_percentile_threshold=
            # "description": "The percentile of cosine dissimilarity that must be exceeded between a group of sentences and the next to form a node.
            # The smaller this number is, the more nodes will be generated",

            semantic_splitter: SemanticSplitterNodeParser = SemanticSplitterNodeParser.from_defaults(
                include_metadata=True,
                buffer_size=40,
                breakpoint_percentile_threshold=90,
                embed_model=embed_model,
                id_func=partial(my_id_func, "semantic-splitter"),
                # sentence_splitter=my_sentence_splitter_callable
            )

            semantic_nodes = semantic_splitter.get_nodes_from_documents(documents=[doc], show_progress=True)

    return semantic_nodes

    # splitter = SentenceSplitter(
    #     chunk_size=1024,
    #     chunk_overlap=20,
    # )
    # nodes = splitter.get_nodes_from_documents(documents)
    # nodes[0]

    # from llama_index.embeddings.ollama import OllamaEmbedding
    #
    # ollama_embedding = OllamaEmbedding(
    #     model_name="llama2",
    #     base_url="http://localhost:11434",
    #     ollama_additional_kwargs={"mirostat": 0},
    # )
    #
    # pass_embedding = ollama_embedding.get_text_embedding_batch(
    #     ["This is a passage!", "This is another passage"], show_progress=True
    # )
    # logger.debug(pass_embedding)
    #
    # query_embedding = ollama_embedding.get_query_embedding("Where is blue?")
    # logger.debug(query_embedding)


#
#
#
#     from llama_index.core.node_parser import SemanticSplitterNodeParser
#     from llama_index.embeddings.openai import OpenAIEmbedding
#
#     embed_model = OpenAIEmbedding()
#     splitter = SemanticSplitterNodeParser(
#         buffer_size=1, breakpoint_percentile_threshold=95, embed_model=embed_model
#     )
#
#
#     from llama_index.llms.ollama import Ollama
#     llm = Ollama(model="llama3.1:latest", request_timeout=120.0)
#     resp = llm.complete("Who is Paul Graham?")
#     logger.debug(resp)
#
#
#
#     from llama_index.core.llms import ChatMessage
#
#     messages = [
#         ChatMessage(
#             role="system", content="You are a pirate with a colorful personality"
#         ),
#         ChatMessage(role="user", content="What is your name"),
#     ]
#     resp = llm.chat(messages)
#
#
#     from llama_index.core import VectorStoreIndex
#     from llama_index.core import PromptTemplate
#     from llama_index.readers.file import PyMuPDFReader
#     loader = PyMuPDFReader()
#     documents = loader.load(file_path="./data/llama2.pdf")
#
#     from llama_index.core import VectorStoreIndex
#     from llama_index.llms.openai import OpenAI
#
#     gpt35_llm = OpenAI(model="gpt-3.5-turbo")
#     gpt4_llm = OpenAI(model="gpt-4")
#
#     index = VectorStoreIndex.from_documents(documents)
#
#     query_str = "What are the potential risks associated with the use of Llama 2 as mentioned in the context?"
#
#     query_engine = index.as_query_engine(similarity_top_k=2, llm=gpt35_llm)
#
#     # use this for testing
#     vector_retriever = index.as_retriever(similarity_top_k=2)
#     response = query_engine.query(query_str)
#     print(str(response))
#
#     def display_prompt_dict(prompts_dict):
#         for k, p in prompts_dict.items():
#             text_md = f"**Prompt Key**: {k}<br>" f"**Text:** <br>"
#             display(Markdown(text_md))
#             print(p.get_template())
#             display(Markdown("<br><br>"))
#
#     prompts_dict = query_engine.get_prompts()
#     display_prompt_dict(prompts_dict)


class MyFileTypes(StrEnum):
    docx = auto()
    md = auto()
    doc = auto()
    xls = auto()
    xlsx = auto()
    yaml = auto()
    html = auto()
    pdf = auto()
    txt = auto()
    any = auto()


def get_files_by_ext_from_dir(filedir: Path, ext: MyFileTypes, maxret: int = -1) -> None | Path | list[Path]:
    files: list[str] = glob.glob("*" if ext == ext.any else f"*.{ext.value}", root_dir=filedir)

    if len(files) == 0:
        return None
    elif len(files) == 1 or maxret == 1:
        return Path(filedir, files[0])

    return [Path(filedir, k) for num, k in enumerate(files) if num < maxret or maxret < 0]


def yaml_raw_file_to_yaml_file(yaml_raw_file: Path, file_base_name: str, title: str) -> Path:
    outf: Path = Path(yaml_raw_file.parent.resolve(), file_base_name + "." + MyFileTypes.yaml.value)  # type: ignore
    if outf.exists():
        logger.debug(f"EXISTS: {outf.resolve()}")
        return outf

    adi: ArleyDocumentInformation = ArleyDocumentInformation.from_xls_converted_yaml_file(
        yaml_raw_file, title=title, doctype=DocTypeEnum.contract.value  # type: ignore
    )

    yaml_data: dict = adi.model_dump(mode="json", by_alias=True, exclude_none=True)

    logger.debug(Helper.get_dict_as_yaml_str(yaml_data))

    with open(outf, "w") as outfile:
        yaml = ruamel.yaml.YAML()
        yaml.indent(sequence=4, offset=2)

        # yaml.dump(xlsx_data, outfile, indent=4, allow_unicode=True, sort_keys=False, default_flow_style=False)
        yaml.dump(yaml_data, outfile)

    return outf


def main(
    docdir: Path,
    plaintxt_based: bool = True,
    llm_md_based_if_not_plaintxt_based: bool = True,
    directly_convert_to_txt: bool = True,
    delete_other_llm_base: bool = True,
) -> int:
    if not docdir.exists() or not docdir.is_dir():
        logger.error(f"{docdir} does not exist or is not a directory")
        return 123

    ret_html: int
    ret_md: int
    ret_pdf: int

    pdf_file: Path
    md_file: Path
    md_llm_file: Path | None
    html_file: Path
    yaml_file: Path
    yaml_raw_file: Path | None
    llm_summary_file: Path | None
    llm_outline_file: Path | None

    docx_file: None | Path | list[Path] = get_files_by_ext_from_dir(
        filedir=docdir, ext=MyFileTypes.docx, maxret=1
    )  # must only be ONE in that dir

    assert isinstance(docx_file, Path)

    if not docx_file or not docx_file.is_file():
        raise Exception(f"Invalid file {docx_file=}")

    file_basename: str = docx_file.name[0 : docx_file.name.rfind(".")]  # type: ignore

    xlsx_file: Path = Path(docx_file.parent, f"{file_basename}.{MyFileTypes.xlsx.value}")
    if not xlsx_file or not xlsx_file.is_file():
        raise Exception(f"Invalid file {xlsx_file=}")

    yaml_raw_file = xlsx_to_yaml_raw(xlsx_file)

    assert yaml_raw_file is not None

    yaml_file = yaml_raw_file_to_yaml_file(
        yaml_raw_file=yaml_raw_file, file_base_name=file_basename, title=file_basename
    )

    adi: ArleyDocumentInformation | None = ArleyDocumentInformation.from_yaml_model_file(yaml_file)
    lang: Literal["en", "de"] = adi.lang.value  # type: ignore

    ret_pdf, pdf_file = docx_to_pdf_via_subprocess(docx_file)  # type: ignore
    ret_html, html_file = docx_to_html_via_subprocess(docx_file)  # type: ignore
    # ret_md, md_file = html_to_md_via_subprocess(html_file)
    ret_md, md_file = docx_to_md_via_subprocess(docx_file)  # type: ignore

    txt_file: Path
    if directly_convert_to_txt:
        txt_ret, txt_file = docx_to_txt_via_subprocess(docx_file)  # type: ignore
    else:
        txt: str | None = parse_markdown(file=md_file, remove_leading_comment_indent=True)
        txt_file = Path(docdir, file_basename + "." + MyFileTypes.txt)
        if txt is not None:
            with open(txt_file, "w") as outfile:
                outfile.write(txt)
                outfile.flush()

    base_file: Path = md_file

    md_llm_file = llm_remove_unnecessary_newlines_and_generate_semantic_markup_file(txt_file)

    assert md_llm_file is not None

    if llm_md_based_if_not_plaintxt_based and not plaintxt_based:
        base_file = md_llm_file

    if plaintxt_based:
        base_file = txt_file

    llm_summary_file = llm_create_summary_file(
        md_or_plaintxt_file=base_file, is_plaintext=plaintxt_based, delete_other_llm_base=delete_other_llm_base
    )
    llm_outline_file = llm_create_outline_file(
        md_or_plaintxt_file=base_file, is_plaintext=plaintxt_based, delete_other_llm_base=delete_other_llm_base
    )

    excerpts: list[Path] | None = create_excerpt_files(
        md_or_plaintxt_file=base_file,
        is_plaintext=plaintxt_based,
        lang=lang,
        do_simple_md_split=False,
        spacy_mode=False,
    )

    if excerpts is not None:
        for exc in excerpts:
            logger.debug(f"\tExcerpt Written to: {exc.absolute()}")

    logger.debug(f"Written to [{ret_pdf=}]: {pdf_file.absolute()}")
    logger.debug(f"Written to [{ret_html=}]: {html_file.absolute()}")
    logger.debug(f"Written to [{ret_md=}]: {md_file.absolute()}")
    logger.debug(f"Written to: {yaml_raw_file.absolute()}")
    logger.debug(f"Written to: {yaml_file.absolute()}")
    logger.debug(f"Written to: {txt_file.absolute()}")
    logger.debug(f"Written to: {md_llm_file.absolute()}")

    assert llm_summary_file is not None
    logger.debug(f"Written to: {llm_summary_file.absolute()}")

    assert llm_outline_file is not None
    logger.debug(f"Written to: {llm_outline_file.absolute()}")

    assert excerpts is not None and adi is not None

    importtovectorstore(
        adi=adi,
        md_or_plaintxt_file=base_file,
        is_plaintext=plaintxt_based,
        llm_summary_file=llm_summary_file,
        llm_outline_file=llm_outline_file,
        excerpts=excerpts,
    )

    return 0


def importdirtovectorstore(
    basedir: Path, plaintext_based: bool, llm_md_based_if_not_plaintxt_based: bool = True
) -> int:
    if not basedir.exists() or not basedir.is_dir():
        raise RuntimeError(f"{basedir.resolve()} does not exist or is not a directory")

    adi: ArleyDocumentInformation | None = None

    txt_base_file: Path | None = None
    md_base_file: Path | None = None
    llm_summary_file: Path | None = None
    llm_outline_file: Path | None = None

    for mefile in basedir.glob("*"):
        logger.debug(f"{mefile.resolve()}")
        if mefile.name.endswith("_llm.md"):
            if llm_md_based_if_not_plaintxt_based:
                md_base_file = mefile
        elif mefile.name.endswith(".md"):
            if not llm_md_based_if_not_plaintxt_based:
                md_base_file = mefile
        elif mefile.name.endswith(".txt"):
            txt_base_file = mefile
        elif mefile.name.endswith("_llm_summary.md") or mefile.name.endswith("_llm_summary.txt"):
            llm_summary_file = mefile
        elif mefile.name.endswith("_llm_outline.md") or mefile.name.endswith("_llm_outline.txt"):
            llm_outline_file = mefile
        elif mefile.name.endswith("_raw.yaml"):
            continue
        elif mefile.name.endswith(".yaml"):
            adi = ArleyDocumentInformation.from_yaml_model_file(mefile)

    assert md_base_file is not None and txt_base_file is not None

    base_file: Path = md_base_file
    if plaintext_based:
        base_file = txt_base_file

    excerpts: list[Path] = []
    llm_excerpt_dir: Path = Path(basedir, "llm_excerpts")

    for mefile in llm_excerpt_dir.glob("*.txt"):
        excerpts.append(mefile)

    if not llm_summary_file or not llm_outline_file or not llm_summary_file or not adi or not base_file:
        raise RuntimeError(f"invalid data {adi==None=} {llm_summary_file=} {llm_outline_file=} {base_file=}")

    importtovectorstore(
        adi=adi,
        md_or_plaintxt_file=base_file,
        is_plaintext=plaintext_based,
        llm_summary_file=llm_summary_file,
        llm_outline_file=llm_outline_file,
        excerpts=excerpts,
    )

    return 0


def importtovectorstore(
    adi: ArleyDocumentInformation,
    md_or_plaintxt_file: Path,
    is_plaintext: bool,
    llm_summary_file: Path,
    llm_outline_file: Path,
    excerpts: list[Path],
    collectionname: str = _CHROMADB_DEFAULT_COLLECTION_NAME,
) -> int:
    cdbconnection: ChromaDBConnection = ChromaDBConnection.get_instance()
    cdbcollection: ChromaCollection = cdbconnection.get_or_create_collection(collectionname)

    fsmap: dict[str, Path] = {"file": md_or_plaintxt_file, "summary": llm_summary_file, "outline": llm_outline_file}

    parent_id_for_excerpts: str = str(adi.id)
    parent_file_name_for_excerpts: str = md_or_plaintxt_file.name
    parent_full_path_for_excerpts: str = str(md_or_plaintxt_file.resolve())

    for mefile in excerpts:
        fsmap[str(mefile.resolve())] = mefile

    for filename in fsmap:  # .keys():
        my_id: str = str(adi.id)

        mefile = fsmap[filename]
        md5: str = Helper.get_md5_for_file(mefile)
        metadata: dict = adi.model_dump(mode="json", by_alias=True, exclude_none=True)
        metadata["md5"] = md5

        metadata["file_name"] = mefile.name
        metadata["full_path"] = str(mefile.resolve())

        metadata = Helper.flatten(metadata)
        metadata = Helper.flatten_lists(metadata)

        is_main_file: bool = False

        match filename:
            case "file":
                is_main_file = True
            case "outline":
                metadata["doctype"] = DocTypeEnum.contract_outline.value
                my_id = f"{my_id}_outline"
                metadata["parent_full_path"] = parent_full_path_for_excerpts
                metadata["parent_file_name"] = parent_file_name_for_excerpts
                metadata["parent_id"] = parent_id_for_excerpts
            case "summary":
                metadata["doctype"] = DocTypeEnum.contract_summary.value
                my_id = f"{my_id}_summary"
                metadata["parent_full_path"] = parent_full_path_for_excerpts
                metadata["parent_file_name"] = parent_file_name_for_excerpts
                metadata["parent_id"] = parent_id_for_excerpts
            case _:
                metadata["doctype"] = DocTypeEnum.contract_excerpt.value
                metadata["parent_full_path"] = parent_full_path_for_excerpts
                metadata["parent_file_name"] = parent_file_name_for_excerpts
                metadata["parent_id"] = parent_id_for_excerpts

                range_str: str = mefile.name[mefile.name.rfind("_") + 1 :]
                range_str = range_str[0 : range_str.rfind(".")]

                logger.debug(f"{mefile.name=} {range_str=}")

                range_from: int = int(range_str[0 : range_str.rfind("-")])
                range_to: int = int(range_str[range_str.rfind("-") + 1 :])

                my_id = f"{my_id}_excerpt_{range_str}"

        metadata["id"] = my_id

        with open(mefile, "r") as infile:
            txt: str = infile.read()

            logger.debug(Helper.get_pretty_dict_json_no_sort(metadata))
            logger.debug(txt)

            cdbconnection.add_document(cdbcollection=cdbcollection, doc_id=my_id, document=txt, metadata=metadata)

        if is_main_file:
            metadata["parent_full_path"] = parent_full_path_for_excerpts
            metadata["parent_file_name"] = parent_file_name_for_excerpts
            metadata["parent_id"] = parent_id_for_excerpts

            metadata["doctype"] = DocTypeEnum.categorization_lookup.value
            my_id = f"{str(adi.id)}_catlookup"

            cdbconnection.add_document(
                cdbcollection=cdbcollection, doc_id=my_id, document=metadata["categorization_tags"], metadata=metadata
            )

            if metadata.get("categorization_targeted_by_prompts"):
                metadata["doctype"] = DocTypeEnum.targeted_by_prompts_lookup.value
                my_id = f"{str(adi.id)}_prompttarget"
                cdbconnection.add_document(
                    cdbcollection=cdbcollection,
                    doc_id=my_id,
                    document=metadata["categorization_targeted_by_prompts"],
                    metadata=metadata,
                )

            metadata["doctype"] = DocTypeEnum.title_lookup.value
            my_id = f"{str(adi.id)}_titlelookup"
            cdbconnection.add_document(
                cdbcollection=cdbcollection, doc_id=my_id, document=metadata["categorization_titles"], metadata=metadata
            )

    return 0


def main_cli() -> int:
    plaintext_based: bool = False
    llm_md_based_if_not_plaintxt_based: bool = True
    directly_convert_to_txt: bool = True

    do_simple_md_split: bool = False
    spacy_mode: bool = False
    delete_old_excerpts: bool = True
    delete_other_llm_base: bool = False

    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        prog="importhelper.py"
    )  # , usage='%(prog)s cmd [options]')

    subparsers = parser.add_subparsers(dest="cmd")
    subparsers.required = True

    #  subparser for convert
    parser_convert = subparsers.add_parser("convert")
    # add a required argument
    parser_convert.add_argument("dir", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster")))

    parser_parsemd = subparsers.add_parser("parsemd")
    # add a required argument
    parser_parsemd.add_argument("file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA.md")))

    parser_parsemd = subparsers.add_parser("parsepdf")
    # add a required argument
    parser_parsemd.add_argument(
        "file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA.pdf"))
    )

    parser_llmparse = subparsers.add_parser("llmparse")
    # add a required argument
    parser_llmparse.add_argument(
        "file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA.txt"))
    )

    parser_llmsummary = subparsers.add_parser("llmsummary")
    # add a required argument
    parser_llmsummary.add_argument(
        "file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA_llm.md"))
    )

    parser_llmoutline = subparsers.add_parser("llmoutline")
    # add a required argument
    parser_llmoutline.add_argument(
        "file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA_llm.md"))
    )

    parser_parsemdsemantically = subparsers.add_parser("parsemdsemantically")
    # add a required argument
    parser_parsemdsemantically.add_argument(
        "file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA_llm.md"))
    )

    parser_createexcerptfiles = subparsers.add_parser("createexcerptfiles")
    # add a required argument
    parser_createexcerptfiles.add_argument(
        "file", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster/NDA_llm.md"))
    )

    parser_import = subparsers.add_parser("importdirtovectorstore")
    # add a required argument
    parser_import.add_argument("dir", type=str, nargs="?", default=str(Path(Path.home(), "ki_vertragsmuster")))

    logger.debug(sys.argv)
    ns: argparse.Namespace = parser.parse_args(sys.argv[1:])

    _ret: int = -1
    mdfile: Path

    if ns.cmd == "convert":
        docdir: Path = Path(ns.dir)
        _ret = main(
            docdir=docdir,
            plaintxt_based=plaintext_based,
            directly_convert_to_txt=directly_convert_to_txt,
            llm_md_based_if_not_plaintxt_based=llm_md_based_if_not_plaintxt_based,
            delete_other_llm_base=delete_other_llm_base,
        )
    elif ns.cmd == "parsepdf":
        pdffile: Path = Path(ns.file)
        _ret = parse_pdf(pdffile)
    elif ns.cmd == "parsemd":
        mdfile = Path(ns.file)
        _ret = 0 if parse_markdown(mdfile) else 234
    elif ns.cmd == "parsemdsemantically":
        mdfile = Path(ns.file)
        _ret = (
            0
            if parse_markdown_or_plaintext_semantically(
                md_or_plaintext_file=mdfile, is_plaintext=mdfile.name.endswith("txt"), lang=None
            )
            else 789
        )
    elif ns.cmd == "createexcerptfiles":
        mdfile = Path(ns.file)
        _ret = (
            0
            if create_excerpt_files(
                md_or_plaintxt_file=mdfile,
                is_plaintext=mdfile.name.endswith("txt"),
                do_simple_md_split=do_simple_md_split,
                spacy_mode=spacy_mode,
                delete_old_excerpts=delete_old_excerpts,
                lang=None,
            )
            else 456
        )
    elif ns.cmd == "llmparse":
        txtfile: Path = Path(ns.file)
        _ret = 0 if llm_remove_unnecessary_newlines_and_generate_semantic_markup_file(txtfile) else 345
    elif ns.cmd == "llmsummary":
        mdfile = Path(ns.file)
        _ret = 0 if llm_create_summary_file(mdfile, is_plaintext=False) else 101
    elif ns.cmd == "llmoutline":
        mdfile = Path(ns.file)
        _ret = 0 if llm_create_outline_file(mdfile, is_plaintext=False) else 910
    elif ns.cmd == "importdirtovectorstore":
        docdir = Path(ns.dir)
        _ret = (
            0
            if importdirtovectorstore(
                basedir=docdir,
                plaintext_based=plaintext_based,
                llm_md_based_if_not_plaintxt_based=llm_md_based_if_not_plaintxt_based,
            )
            else 111
        )

    return _ret


if __name__ == "__main__":
    _OLLAMA_DEFAULT_MODEL = "hermes3:8b-llama3.1-fp16"  # "nous-hermes2-mixtral:8x7b"  # hermes3:70b-llama3.1-q4_0"  # "hermes3:latest"  # hermes3:70b-llama3.1-q4_0"  # "nous-hermes2-mixtral:8x7b"

    exit(main_cli())

    # https://stackoverflow.com/questions/25626109/python-argparse-conditionally-required-arguments/70716254#70716254

    # import argparse
    #
    # # First parse the deciding arguments.
    # deciding_args_parser = argparse.ArgumentParser(add_help=False)
    # deciding_args_parser.add_argument(
    #         '--argument', required=False, action='store_true')
    # deciding_args, _ = deciding_args_parser.parse_known_args()
    #
    # # Create the main parser with the knowledge of the deciding arguments.
    # parser = argparse.ArgumentParser(
    #         description='...', parents=[deciding_args_parser])
    # parser.add_argument('-a', required=deciding_args.argument)
    # parser.add_argument('-b', required=deciding_args.argument)
    # arguments = parser.parse_args()
    #
    # print(arguments)

    # # Add a checker to set the required flags
    # checker = argparse.ArgumentParser()
    # checker.add_argument("--server-1", action="store_true")
    # checker.add_argument("--server-2", action="store_true")
    # checks, _ = checker.parse_known_args()
    #
    # # Create a new main parser, inheriting the argument definitions
    # #   from the checker, to not duplicate definitions.
    # parser = argparse.ArgumentParser(parents=[checker])
    #
    # # Use the flags from `checker` to flip the requirement state
    # #   as needed
    # parser.add_argument("--config-1", required=checks.server_1)
    # parser.add_argument("--config-2", required=checks.server_2)
    #
    # parsed_args = parser.parse_args()
